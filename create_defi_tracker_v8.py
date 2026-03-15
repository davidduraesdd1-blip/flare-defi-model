from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

# Config dictionary for easier updates
CONFIG = {
    'save_path': r"C:\Users\dduraes\OneDrive - Cherry Creek School District\Desktop\SuperGrok Mathematically Model\DeFi_Tracker_Model_v8.xlsx",
    'positions': [
        {
            'id': 35399,
            'pool': "WFLR - USD₮0",
            'liquidity': 14790,
            'flr_balance': "231.16 FLR",
            'usdt_balance': "0.05441 USD₮0",
            'fees': 32.74,
            'rewards': "11,640 RFLR",
            'est_value': 14816,
            'entry_value': "",  # User fills  
            'il': "",  # User fills
            'actual_roi': "",  # User fills
        },
        {
            'id': 36910,
            'pool': "FXRP - WFLR",
            'liquidity': 1130,
            'flr_balance': "231.16 WFLR",
            'usdt_balance': "0 FXRP",
            'fees': 0.04,
            'rewards': "0.04 RFLR",
            'est_value': 1134,
            'entry_value': "",
            'il': "",
            'actual_roi': "",
        },
    ],
    'model_data': [
        ["WFLR - USD₮0", 35, 5250, "99.27%", "~133%", "🟡", "4/10"],
        ["FXRP - USD₮0", 25, 3750, "106.27%", "~142%", "🟡", "5/10"],
        ["USDT0 lending/vaults", 20, 3000, "15.3–16.9%", "~16–18%", "🟢", "9/10"],
        ["sFLR - WFLR", 10, 1500, "31.45%", "~37%", "🟢", "7/10"],
        ["WFLR - FXRP", 5, 750, "109.97%", "~148%", "🔴", "4/10"],
        ["HLN - FXRP", 5, 750, "125.47%", "~168%", "🔴", "3/10"]
    ],
    'summary': {
        'projected_return': "26500–39500",
        'blended_apy': "210–265%",
        'diversification': "4/10",
    },
}

# Load or create workbook
try:
    wb = load_workbook(CONFIG['save_path'])
except FileNotFoundError:
    wb = Workbook()

# Sheet 1: Tracker Holdings
ws1 = wb.get_sheet_by_name("Tracker Holdings") if "Tracker Holdings" in wb.sheetnames else wb.active
ws1.title = "Tracker Holdings"

# Headers (expanded for Performance Tracker)
headers = ["Position ID", "Pool", "Liquidity Value ($)", "FLR Balance", "USD₮0 Balance", "Unclaimed Fees ($)", "Incentive Rewards", "Current Est. Value ($)", "Entry Value ($)", "ROI (%)", "Est. IL (%)", "Actual ROI (%)", "Variance (%)", "Performance Alert"]
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Load positions
for idx, position in enumerate(CONFIG['positions'], start=2):
    ws1.cell(row=idx, column=1, value=position['id'])
    ws1.cell(row=idx, column=2, value=position['pool'])
    ws1.cell(row=idx, column=3, value=position['liquidity'])
    ws1.cell(row=idx, column=4, value=position['flr_balance'])
    ws1.cell(row=idx, column=5, value=position['usdt_balance'])
    ws1.cell(row=idx, column=6, value=position['fees'])
    ws1.cell(row=idx, column=7, value=position['rewards'])
    ws1.cell(row=idx, column=8, value=position['est_value'])
    ws1.cell(row=idx, column=9, value=position['entry_value'])
    ws1.cell(row=idx, column=10, value="=IF(I{0}=0,\"N/A\",(H{0}-I{0})/I{0})".format(idx))
    ws1.cell(row=idx, column=11, value=position['il'])
    ws1.cell(row=idx, column=12, value=position['actual_roi'])
    ws1.cell(row=idx, column=13, value="=IF(L{0}=\"N/A\",\"N/A\",J{0}-L{0})".format(idx))
    ws1.cell(row=idx, column=14, value="=IF(M{0}=\"N/A\",\"N/A\",IF(M{0}>5,\"Green: Good\",IF(M{0}>-5,\"Yellow: Okay\",\"Red: Underperforming\")))".format(idx))

# Totals
ws1.cell(row=4, column=1, value="Total")
ws1.cell(row=4, column=3, value="=SUM(C2:C3)")
ws1.cell(row=4, column=6, value="=SUM(F2:F3)")
ws1.cell(row=4, column=8, value="=SUM(H2:H3)")

# Green fill for positive unclaimed fees
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ws1['F2'].fill = green_fill
ws1['F3'].fill = green_fill

# Auto column widths
for col in range(1, 15):
    ws1.column_dimensions[get_column_letter(col)].width = 18

# Sheet 2: Model Allocation
ws2 = wb.get_sheet_by_name("Model Allocation") if "Model Allocation" in wb.sheetnames else wb.create_sheet("Model Allocation")
ws2['A1'] = "Pool"
ws2['B1'] = "%"
ws2['C1'] = "$ Amount"
ws2['D1'] = "Nominal APR"
ws2['E1'] = "Effective APY"
ws2['F1'] = "Risk"
ws2['G1'] = "Sustainability"

model_data = CONFIG['model_data']
for row_idx, row_data in enumerate(model_data, start=2):
    for col_idx, value in enumerate(row_data, 1):
        ws2.cell(row=row_idx, column=col_idx, value=value)

ws2['A8'] = "Total %"
ws2['B8'] = "=SUM(B2:B7)"

# Sheet 3: Summary
ws3 = wb.get_sheet_by_name("Summary") if "Summary" in wb.sheetnames else wb.create_sheet("Summary")
ws3['A1'] = "Total Tracker Value"
ws3['B1'] = "= 'Tracker Holdings'!H4"
ws3['A2'] = "Projected Model Return (mid)"
ws3['B2'] = CONFIG['summary']['projected_return']
ws3['A3'] = "Blended APY"
ws3['B3'] = CONFIG['summary']['blended_apy']
ws3['A4'] = "Diversification Score"
ws3['B4'] = CONFIG['summary']['diversification']

# Sheet 4: Weekly Log
ws4 = wb.get_sheet_by_name("Weekly Log") if "Weekly Log" in wb.sheetnames else wb.create_sheet("Weekly Log")
ws4['A1'] = "Date"
ws4['B1'] = "Total Value"
ws4['C1'] = "Unclaimed Fees"
ws4['D1'] = "Notes"

# New Sheet: Performance Chart
ws5 = wb.get_sheet_by_name("Performance Chart") if "Performance Chart" in wb.sheetnames else wb.create_sheet("Performance Chart")
ws5['A1'] = "Date"
ws5['B1'] = "Projected ROI (%)"
ws5['C1'] = "Actual ROI (%)"

# Example data (you fill in more rows over time)
ws5['A2'] = "2/17/2026"
ws5['B2'] = 200  # Example projected
ws5['C2'] = 185  # Example actual (user enters)

# Create line chart
chart = LineChart()
chart.title = "Projected vs Actual ROI"
chart.y_axis.title = "ROI (%)"
chart.x_axis.title = "Date"

data = Reference(ws5, min_col=2, min_row=1, max_col=3, max_row=10)  # Adjust max_row as you add data
cats = Reference(ws5, min_col=1, min_row=2, max_row=10)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws5.add_chart(chart, "E2")

# Save
wb.save(CONFIG['save_path'])

print("Updated Excel file v8 created successfully with optimizations!")

# Basic Test Function
def test_script():
    try:
        wb = load_workbook(CONFIG['save_path'])
        ws1 = wb["Tracker Holdings"]
        assert ws1['H4'].value == 25950, "Total value sum failed"
        print("Test passed: Total value sum is correct")
    except AssertionError as e:
        print("Test failed: ", e)

test_script()